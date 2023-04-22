import openpyxl

from collections import defaultdict

my_path = r"C:\Users\footb\OneDrive\Рабочий стол\Учеба\4 семестр\Теория принятий решений (4 семестр)\Таблицы\table.xlsx"
wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = wb_obj.active
Arr1 = []
Arr2 = []
Arr3 = []
Arr4 = []
Arr5 = []
Arr6 = []
main1 = [[0 for x in range(10)] for y in range(10)]
chet = my_sheet_obj.max_row
for i in range(0, chet - 2):
    main1[i][i] = "X"
for i in range(2, chet):
    Arr1.append(my_sheet_obj.cell(i, 3).value)  # Arr.append записывает строку в список Arr (столбец 1)
    Arr2.append(my_sheet_obj.cell(i, 4).value)  # Столбец 2
    Arr3.append(my_sheet_obj.cell(i, 5).value)
    Arr4.append(my_sheet_obj.cell(i, 6).value)
    Arr5.append(my_sheet_obj.cell(i, 7).value)  # Столбец 5
    Arr6.append(my_sheet_obj.cell(i, 8).value)  # Столбец 6


def Electra(porog):
    for i in range(0, chet - 2):
        a = Arr1[i]  # Записываем первый элемент 1/2/3/4/5/6/7/8/9/10 строки
        b = Arr2[i]
        c = Arr3[i]
        d = Arr4[i]
        e = Arr5[i]
        r = Arr6[i]  # Записываем шестой элемент 1/2/3/4/5/6/7/8/9/10 строки
        for j in range(i + 1, chet - 2):
            a1 = Arr1[j]  # Записываем первый элемент 2/3/4/5/6/7/8/9/10 строки
            b1 = Arr2[j]
            c1 = Arr3[j]
            d1 = Arr4[j]
            e1 = Arr5[j]
            r1 = Arr6[j]  # Записываем шестой элемент 2/3/4/5/6/7/8/9/10 строки
            p = []
            n = []
            if a == a1:
                p.append(0)
                n.append(0)
            elif a < a1:
                p.append(0)
                n.append(2)
            elif a > a1:
                p.append(2)
                n.append(0)
            if b == b1:
                p.append(0)
                n.append(0)
            elif b < b1:
                p.append(3)
                n.append(0)
            elif b > b1:
                p.append(0)
                n.append(3)
            if c == c1:
                p.append(0)
                n.append(0)
            elif c < c1:
                p.append(5)
                n.append(0)
            elif c > c1:
                p.append(0)
                n.append(5)
            if d == d1:
                p.append(0)
                n.append(0)
            elif d < d1:
                p.append(5)
                n.append(0)
            elif d > d1:
                p.append(0)
                n.append(5)
            if e == e1:
                p.append(0)
                n.append(0)
            elif e < e1:
                p.append(0)
                n.append(4)
            elif e > e1:
                p.append(4)
                n.append(0)
            if r == r1:
                p.append(0)
                n.append(0)
            elif r < r1:
                p.append(0)
                n.append(2)
            elif r > r1:
                p.append(2)
                n.append(0)
            P = sum(p)
            N = sum(n)
            if P == 0:
                main1[i][j] = "-"
                main1[j][i] = float('inf')
            elif N == 0:
                main1[j][i] = "-"
                main1[i][j] = float('inf')
            else:
                Dij = P / N
                Dji = N / P
                if Dij > 1:
                    main1[i][j] = float('{:.2f}'.format(Dij))
                if Dji > 1:
                    main1[j][i] = float('{:.2f}'.format(Dji))
    for i in range(0, chet - 2):  # Заполнение таблицы с учётом порога
        for j in range(0, chet - 2):
            if type(main1[i][j]) == str:
                continue
            if main1[i][j] < porog:
                main1[i][j] = "-"
    Vivod = [[str(e) for e in row] for row in main1]  # 115-119 вывод таблицы
    lens = [max(map(len, col)) for col in zip(*Vivod)]
    fmt = '\t'.join('{{:{}}}'.format(x) for x in lens)
    table = [fmt.format(*row) for row in Vivod]
    print('\n'.join(table))


def get_graph(main1, porog):  # Получаем граф предпочтений
    tab = defaultdict(list)
    V = len(main1)
    for i in range(0, len(main1)):
        for j in range(0, len(main1[0])):
            if main1[i][j] == 'X' or main1[i][j] == '-':
                continue
            if main1[i][j] > porog:
                tab[i + 1].append(j + 1)
    return tab


def print_graph(G):  # Красивый вывод графа предпочтений
    for node, sub_nodes in G.items():
        print(str(node) + ": [", end="")
        for sub_node in sub_nodes:
            print(str(sub_node) + " ", end="")
        print("]")


Vvod = 1
while Vvod != 0:
    print('Меню:')
    print('1 - Вывести матрицу предпочтений')
    print('2 - Установить порог и вывести матрицу предпочтений')
    print('0 - Завершение работы программы')
    Vvod = int(input())
    if Vvod == 1:
        Electra(0)
        tab = get_graph(main1, 0)
        print_graph(tab)
    elif Vvod == 2:
        print('Установите порог: ')
        porog = int(input())
        Electra(porog)
        tab = get_graph(main1, porog)
        print_graph(tab)
    elif Vvod == 0:
        print('Работы программы завершена успешно, спасибо!')
    else:
        print('Ошибка!')
